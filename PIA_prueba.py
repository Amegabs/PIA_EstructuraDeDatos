import datetime
import csv
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
from sqlite3 import Error
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)


def verificar_estado_inicial():
    """Verifica si existe un estado previo de la base de datos y muestra mensaje al usuario"""
    base_datos = "coworking.db"

    if not os.path.exists(base_datos):
        print(
            "No se encontró una versión anterior. Iniciando con un estado inicial vacío."
        )
        return False
    else:
        try:
            with sqlite3.connect(base_datos) as conexion:
                cursor = conexion.cursor()
                cursor.execute("SELECT COUNT(*) FROM Clientes")
                clientes = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM Salas")
                salas = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM Reservaciones")
                reservaciones = cursor.fetchone()[0]

                if clientes == 0 and salas == 0 and reservaciones == 0:
                    print(
                        "No se encontraron datos previos. Iniciando con un estado inicial vacío."
                    )
                    return False
                else:
                    print(
                        "Se encontró un estado previo con datos. Continuando con la sesión."
                    )
                    return True
        except Exception as e:
            print(f"No se pudo verificar el estado previo: {e}")
            return False


def iniciar_bd():
    """Funcion que crea la base de datos y las tablas"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Clientes (id_cliente INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellido TEXT NOT NULL)"""
            )
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Salas (id_sala INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL)"""
            )
            cursor.execute(
                "CREATE TABLE IF NOT EXISTS Turnos (id_turno INTEGER PRIMARY KEY, turno TEXT NOT NULL)"
            )
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Reservaciones (id_reservaciones INTEGER PRIMARY KEY, id_cliente INTEGER NOT NULL,
                 id_sala INTEGER NOT NULL, fecha TEXT, id_turno INTEGER, evento TEXT NOT NULL, estatus TEXT NOT NULL DEFAULT 'Activa',
                   FOREIGN KEY(id_cliente) REFERENCES Clientes(id_cliente), FOREIGN KEY(id_sala) REFERENCES Salas(id_sala), FOREIGN KEY(id_turno) REFERENCES Turnos(id_turno))"""
            )
            cursor.execute("SELECT COUNT(*) FROM Turnos")
            if cursor.fetchone()[0] == 0:
                cursor.executemany(
                    "INSERT INTO Turnos (id_turno, turno) VALUES (?, ?)",
                    [(1, "Matutino"), (2, "Vespertino"), (3, "Nocturno")],
                )
    except Error as e:
        print(e)
    except Exception as e:
        print(f"Error inesperado: {e}")


def pedir_nombres():
    """Funcion que pide el nombre"""
    while True:
        nombre = input("Nombre(s): ")
        if nombre == "":
            return None
        if nombre.replace(" ", "").isalpha():
            return nombre
        print("Error, ingresa el nombre correctamente. Solo letras.")


def pedir_apellidos():
    """Funcion que pide el apellido"""
    while True:
        apellido = input("Apellido(s): ")
        if apellido == "":
            return None
        if apellido.replace(" ", "").isalpha():
            return apellido
        print("Error, Ingresa el apellido correctamente. Solo letras.")


def pedir_cupo():
    """Funcion que pide el cupo de la sala"""
    while True:
        cupo = input("Cupo de la sala: ")
        if cupo == "":
            return None
        if cupo.isdigit():
            return cupo
        print("Error, Ingresa el número correctamente. Solo números.")


def registrar_reservacion():
    """Funcion que registrara una nueva reservacion en alguna sala disponible"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "SELECT id_cliente, nombre, apellido FROM Clientes ORDER BY apellido, nombre"
            )
            clientes = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not clientes:
        print(
            "No hay clientes registrados. Primero debe registrar un cliente para continuar."
        )
        return

    print("*" * 75)
    print(f"**{'CLIENTES REGISTRADOS':^71}**")
    print("*" * 75)
    print("{:<15} {:<30} {:<30} ".format("Clave Cliente", "Nombre(s)", "Apellido(s)"))
    print("*" * 75)
    for cliente in clientes:
        print("{:<15} {:<30} {:<30}".format(cliente[0], cliente[1], cliente[2]))
    else:
        print("*" * 75)

    intentos_cliente = 0
    while True:
        entrada = input("Ingrese la Clave del Cliente: ").strip()
        if entrada == "":
            intentos_cliente += 1
            if intentos_cliente >= 2:
                print("No ingresaste un cliente. Presiona ENTER para regresar al menu.")
                entrada = input("Ingrese la Clave del Cliente: ").strip()
                if entrada == "":
                    return
            else:
                print("No ingresaste un cliente. Intentalo de nuevo.")
                continue

        try:
            id_cliente = int(entrada)
        except ValueError:
            print("Debe ingresar un numero valido")
            continue
        if not any(cliente[0] == id_cliente for cliente in clientes):
            print("Cliente no encontrado. Intente de nuevo.")
            continue
        break

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT id_sala, nombre, cupo FROM Salas")
            salas = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not salas:
        print("No hay salas registradas. Primero registre una sala para continuar.")
        return

    print("*" * 65)
    print(f"**{'SALAS REGISTRADOS':^61}**")
    print("*" * 65)
    print("{:<15} {:<30} {:<20} ".format("Clave Sala", "Nombre Sala", "Cupo"))
    print("*" * 65)
    for sala in salas:
        print("{:<15} {:<30} {:<20}".format(sala[0], sala[1], sala[2]))
    else:
        print("*" * 65)

    intentos_sala = 0
    while True:
        entrada = input("Ingrese la Clave de la sala: ").strip()
        if entrada == "":
            intentos_sala += 1
            if intentos_sala >= 2:
                print("No ingresaste una sala. Presiona ENTER para regresar al menu.")
                entrada = input("Ingrese la Clave de la sala: ").strip()
                if entrada == "":
                    return
            else:
                print("No ingresaste una sala. Intentalo de nuevo.")
                continue

        try:
            id_sala = int(entrada)
        except ValueError:
            intentos_sala += 1
            print("Debe ingresar un numero valido.")
            continue
        if not any(sala[0] == id_sala for sala in salas):
            print("Sala no encontrada. Intente de nuevo.")
            continue
        break

    print("")
    print("INGRESA LA INFORMACION DE TU RESERVACION")

    intentos_fecha = 0
    while True:
        Fecha_str = input("Ingrese la fecha del evento (MM-DD-AAAA): ").strip()
        if Fecha_str == "":
            intentos_fecha += 1
            if intentos_fecha >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                Fecha_str = input("Ingrese la fecha del evento (MM-DD-AAAA): ").strip()
                if Fecha_str == "":
                    return
            else:
                print("No ingresaste una fecha. Intentelo de nuevo.")
                continue

        try:
            Hoy = datetime.date.today()
            Fecha_evento = datetime.datetime.strptime(Fecha_str, "%m-%d-%Y").date()
            fecha_sql = Fecha_evento.strftime("%Y-%m-%d")
            FechaAnticipada = (Fecha_evento - Hoy).days

            if FechaAnticipada < 2:
                intentos_fecha += 1
                if intentos_fecha >= 2:
                    print(
                        "La reservacion debe ser mayor a 2 dias de anticipacion. Intentalo nuevamente o ENTER para regresar al menu."
                    )
                    Fecha_str = input(
                        "Ingrese la fecha del evento (MM-DD-AAAA): "
                    ).strip()
                    if Fecha_str == "":
                        return
                else:
                    print(
                        "La reservacion debe ser mayor a 2 dias de anticipacion. Intentelo de nuevo."
                    )
                    continue

            if Fecha_evento.weekday() == 6:
                print("No se pueden realizar reservaciones los domingos.")
                lunes_siguiente = Fecha_evento + datetime.timedelta(days=1)
                print(
                    f"Se propone automáticamente el lunes siguiente: {lunes_siguiente.strftime('%m-%d-%Y')}"
                )
                respuesta = input("¿Aceptas esta fecha? (S/N): ").strip().upper()
                if respuesta == "S":
                    Fecha_evento = lunes_siguiente
                else:
                    print("Por favor, ingresa otra fecha que cumpla las condiciones.")
                    continue

            break
        except ValueError:
            intentos_fecha += 1
            if intentos_fecha >= 2:
                print(
                    "Formato de fecha Incorrecto, use MM-DD-AAAA o ENTER para volver al menu."
                )
                Fecha_str = input("Ingrese la fecha del evento (MM-DD-AAAA): ").strip()
                return
            else:
                print(
                    "Formato de fecha incorrecto, use MM-DD-AAAA. Intentelo de nuevo."
                )
                continue

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT id_turno, turno FROM Turnos")
            turnos = cursor.fetchall()
    except Error as e:
        print(e)
        return
    print("*" * 20)
    print(f"{'POSIBLES TURNOS':^20}")
    print("*" * 20)
    for turno in turnos:
        print(f"{turno[0]} : {turno[1]}")
    else:
        print("*" * 20)

    while True:
        turno = input(
            "Selecciona el ID del turno a escoger (ENTER para volver al menu): "
        ).strip()
        if turno == "":
            print("No ingresaste un turno. Volviendo al menu.")
            return
        try:
            id_turno = int(turno)
        except ValueError:
            print("Debe ingresar un numero valido.")
            continue
        if not any(turno[0] == id_turno for turno in turnos):
            print("Turno no encontrado. Intente nuevamente.")
            continue
        try:
            with sqlite3.connect("coworking.db") as conexion:
                cursor = conexion.cursor()
                cursor.execute(
                    "SELECT * FROM Reservaciones WHERE id_sala=? AND fecha=? AND id_turno=?",
                    (id_sala, fecha_sql, id_turno),
                )
                ocupado = cursor.fetchone()
            if ocupado:
                print("Esta sala ya esta reservada en esta fecha y turno. Elija otro.")
                continue
            break
        except Error as e:
            print(e)
            return

    intentos_evento = 0
    while True:
        nombre_evento = input("Ingresa el nombre del evento: ").strip()
        if nombre_evento == "":
            intentos_evento += 1
            if intentos_evento >= 2:
                print(
                    "No ingresaste un nombre de evento valido. Intentalo de nuevo o presiona ENTER para volver al menu."
                )
                nombre_evento = input("Ingresa el nombre del evento: ").strip()
                if nombre_evento == "":
                    return
            else:
                print("No ingresaste un nombre de evento valido. Intentalo de nuevo")
            continue
        if not nombre_evento.replace(" ", "").isalpha():
            print(
                "El nombre del evento solo puede contener letras. Intentalo de nuevo."
            )
            continue
        break

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """
                INSERT INTO Reservaciones (id_cliente, id_sala, fecha, id_turno, evento, estatus)
                VALUES (?, ?, ?, ?, ?, 'Activa')
            """,
                (
                    id_cliente,
                    id_sala,
                    fecha_sql,
                    id_turno,
                    nombre_evento,
                ),
            )
            print("Reservación registrada exitosamente.")
    except Error as e:
        print(e)


def editar_reservacion():
    """Funcion que editara el nombre de la reservacion seleccionada por un rango de fechas"""
    intento_fecha1 = 0
    while True:
        fecha_inicio_str = input("Ingresa la fecha de inicio (MM-DD-AAAA): ").strip()
        if fecha_inicio_str == "":
            intento_fecha1 += 1
            if intento_fecha1 >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                fecha_inicio_str = input(
                    "Ingresa la fecha de inicio (MM-DD-AAAA): "
                ).strip()
                if fecha_inicio_str == "":
                    return
            else:
                print("No ingresaste una fecha. Intentalo de nuevo.")
                continue
        try:
            fecha_inicio = datetime.datetime.strptime(
                fecha_inicio_str, "%m-%d-%Y"
            ).date()
        except ValueError:
            print("Error en el formato de fechas, use MM-DD-AAAA.")
            continue
        break

    intento_fecha2 = 0
    while True:
        fecha_fin_str = input("Ingresa la fecha de fin (MM-DD-AAAA): ").strip()
        if fecha_fin_str == "":
            intento_fecha2 += 1
            if intento_fecha2 >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                fecha_fin_str = input("Ingresa la fecha de fin (MM-DD-AAAA): ").strip()
                if fecha_fin_str == "":
                    return
            else:
                print("No ingresaste una fecha. Intentalo de nuevo.")
                continue
        try:
            fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()
        except ValueError:
            print("Error en el formato de fechas, use MM-DD-AAAA.")
            continue
        break

    fecha_inicio_iso = fecha_inicio.strftime("%Y-%m-%d")
    fecha_fin_iso = fecha_fin.strftime("%Y-%m-%d")

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """SELECT r.id_reservaciones, c.nombre || ' ' || c.apellido AS cliente, r.id_sala, r.fecha, t.turno, r.evento
                    FROM Reservaciones r
                    JOIN Clientes c ON r.id_cliente = c.id_cliente
                    JOIN Turnos t ON r.id_turno = t.id_turno
                    WHERE r.fecha BETWEEN ? AND ?""",
                (fecha_inicio_iso, fecha_fin_iso),
            )
            reservaciones = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not reservaciones:
        print("No se encontraron reservaciones en el rango indicado.")
        return

    print("*" * 104)
    print(f"**{'RESERVACIONES ENCONTRADAS':^100}** ")
    print("*" * 104)
    print(
        "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
            "ID Reserva",
            "Cliente",
            "ID Sala",
            "Fecha",
            "Turno",
            "Nombre evento",
        )
    )
    print("*" * 104)
    for reserva in reservaciones:
        print(
            "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
                reserva[0],
                reserva[1],
                reserva[2],
                reserva[3],
                reserva[4],
                reserva[5],
            )
        )
    else:
        print("*" * 104)

    intento_reservacion = 0
    while True:
        clave = input("\nIngrese la clave de la reservación que desea editar: ").strip()
        if clave == "":
            intento_reservacion += 1
            if intento_reservacion >= 2:
                print(
                    "No ingresaste una reservacion valida. Presiona ENTER para regresar al menu."
                )
                clave = input(
                    "Ingrese la clave de la reservación que desea editar: "
                ).strip()
                if clave == "":
                    return
            else:
                print("No ingresaste una reservacion. Intentalo de nuevo.")
                continue
        try:
            clave_editar = int(clave)
        except ValueError:
            print("Debe ingresar un numero valido")
            continue

        if clave_editar not in [r[0] for r in reservaciones]:
            print("Clave inválida, intente de nuevo.")
            continue
        break

    intentos_nombre = 0
    while True:
        nuevo_nombre = input("Ingrese el nuevo nombre del evento: ").strip()
        if nuevo_nombre == "":
            intentos_nombre += 1
            if intentos_nombre >= 2:
                print(
                    "No ingresaste el nuevo nombre del evento. Presiona ENTER para regresar al menu."
                )
                nuevo_nombre = input("Ingrese el nuevo nombre del evento: ").strip()
                if nuevo_nombre == "":
                    return
            else:
                print("No ingresaste el nuevo nombre. Intentalo de nuevo.")
            continue
        if not nuevo_nombre.replace(" ", "").isalpha():
            print(
                "El nombre del evento solo puede contener letras. Intentalo de nuevo."
            )
            continue
        break

    if not nuevo_nombre:
        print("El nombre del evento no puede quedar vacío.")
        return

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "UPDATE Reservaciones SET evento=? WHERE id_reservaciones=?",
                (nuevo_nombre, clave_editar),
            )
            print("Nombre del evento actualizado con éxito.")
    except Error as e:
        print(f"Error en la base de datos: {e}")


def consultar_reservacion():
    """Funcion que consultara las reservaciones existentes para una fecha especifica"""
    try:
        while True:
            fecha_inicio_str = input("Ingrese la fecha inicial (MM-DD-AAAA): ").strip()
            if fecha_inicio_str == "":
                print("Consulta cancelada.")
                return
            fecha_inicio = datetime.datetime.strptime(
                fecha_inicio_str, "%m-%d-%Y"
            ).date()
            break

        while True:
            fecha_fin_str = input("Ingrese la fecha final (MM-DD-AAAA): ").strip()
            if fecha_fin_str == "":
                print("Consulta cancelada.")
                return
            fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()
            if fecha_fin < fecha_inicio:
                print("La fecha final no puede ser menor que la fecha inicial.")
                continue
            break
    except ValueError:
        print("Formato incorrecto. Use MM-DD-AAAA. Intente nuevamente.")

    fecha_inicio_iso = fecha_inicio.strftime("%Y-%m-%d")
    fecha_fin_iso = fecha_fin.strftime("%Y-%m-%d")

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """SELECT r.id_reservaciones, c.nombre || ' ' || c.apellido AS cliente, r.id_sala, r.fecha, t.turno, r.evento
                            FROM Reservaciones r
                            JOIN Clientes c ON r.id_cliente = c.id_cliente
                            JOIN Turnos t ON r.id_turno = t.id_turno
                            WHERE r.estatus = 'Activa'
                            AND r.fecha BETWEEN ? AND ?""",
                (fecha_inicio_iso, fecha_fin_iso),
            )
            filas = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not filas:
        print("No hay reservaciones registradas.")
        return

    print("*" * 104)
    print(f"**{'RESERVACIONES ENCONTRADAS':^100}**")
    print("*" * 104)
    print(
        "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
            "Id Reserva", "Cliente", "Id Sala", "Fecha", "Turno", "Evento"
        )
    )
    print("*" * 104)

    for fila in filas:
        fecha_evento = datetime.datetime.strptime(fila[3], "%Y-%m-%d").strftime(
            "%m-%d-%Y"
        )
        print(
            "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
                fila[0], fila[1], fila[2], fecha_evento, fila[4], fila[5]
            )
        )
    else:
        print("*" * 104)

    while True:
        try:
            print("\n¿Deseas exportar los datos?")
            print("1. Exportar a CSV")
            print("2. Exportar a Excel")
            print("3. Exportar a JSON")
            print("4. No exportar")
            opcion = input("Selecciona una opción: ")
            if opcion.isdigit():
                if opcion == "1":
                    exportar_csv()
                    break
                elif opcion == "2":
                    exportar_excel()
                    break
                elif opcion == "3":
                    exportar_json()
                    break
                elif opcion == "4":
                    print("No se exportaron los datos.")
                    break
                else:
                    print("Error, ingresa una opcion valida")
            else:
                print("No ingresaste una opcion valida. No se exportaran los datos.")
        except ValueError:
            print("Opcion no valida. No se exportaran los datos.")


def registrar_cliente():
    """Funcion que registrara a un nuevo cliente"""

    intentos_nombre = 0
    while True:
        nombre_cliente = pedir_nombres()
        if nombre_cliente:
            break
        intentos_nombre += 1
        if intentos_nombre >= 2:
            print("No ingresaste un nombre. Presiona ENTER para regresar al menú.")
            nombre_cliente = pedir_nombres()
            break
        else:
            print("No ingresaste un nombre. Intentalo nuevamente.")
    if not nombre_cliente:
        return

    intentos_apellido = 0
    while True:
        apellido_cliente = pedir_apellidos()
        if apellido_cliente:
            break
        intentos_apellido += 1
        if intentos_apellido >= 2:
            print("No ingresaste un apellido. Presiona ENTER para regresar al menú.")
            apellido_cliente = pedir_apellidos()
            break
        else:
            print("No ingresaste un apellido. Intentalo nuevamente.")
    if not apellido_cliente:
        return

    cliente = (nombre_cliente, apellido_cliente)

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "INSERT INTO Clientes (nombre, apellido) VALUES(?,?)", cliente
            )
            print("Cliente agregado exitosamente.")
    except Error as e:
        print(e)


def registrar_sala():
    """Funcion que registrara una nueva sala"""
    intentos_nombre = 0
    while True:
        nombre_sala = pedir_nombres()
        if nombre_sala:
            break
        intentos_nombre += 1
        if intentos_nombre >= 2:
            print(
                "No ingresaste un nombre de sala. Presiona ENTER para regresar al menu."
            )
            nombre_sala = pedir_nombres()
            break
        else:
            print("No ingresaste un nombre de sala. Intentalo nuevamente.")
    if not nombre_sala:
        return

    intentos_cupo = 0
    while True:
        cupo_sala = pedir_cupo()
        if cupo_sala:
            break
        intentos_cupo += 1
        if intentos_cupo >= 2:
            print(
                "No ingresaste un cupo de sala. Presiona ENTER para regresar al menu."
            )
            cupo_sala = pedir_cupo()
            break
        else:
            print("No ingresaste un cupo de sala. Intentalo nuevamente.")
    if not cupo_sala:
        return

    sala = (nombre_sala, cupo_sala)

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO Salas (nombre, cupo) VALUES(?,?)", sala)
            print("Sala creada exitosamente.")
    except Error as e:
        print(e)


def cancelar_reservacion():
    intento_fecha1 = 0
    while True:
        fecha_inicio_str = input("Ingresa la fecha de inicio (MM-DD-AAAA): ").strip()
        if fecha_inicio_str == "":
            intento_fecha1 += 1
            if intento_fecha1 >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                fecha_inicio_str = input(
                    "Ingresa la fecha de inicio (MM-DD-AAAA): "
                ).strip()
                if fecha_inicio_str == "":
                    return
            else:
                print("No ingresaste una fecha. Intentalo de nuevo.")
                continue
        try:
            fecha_inicio = datetime.datetime.strptime(
                fecha_inicio_str, "%m-%d-%Y"
            ).date()
        except ValueError:
            print("Error en el formato de fechas, use MM-DD-AAAA.")
            continue
        break

    intento_fecha2 = 0
    while True:
        fecha_fin_str = input("Ingresa la fecha de fin (MM-DD-AAAA): ").strip()
        if fecha_fin_str == "":
            intento_fecha2 += 1
            if intento_fecha2 >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                fecha_fin_str = input("Ingresa la fecha de fin (MM-DD-AAAA): ").strip()
                if fecha_fin_str == "":
                    return
            else:
                print("No ingresaste una fecha. Intentalo de nuevo.")
                continue
        try:
            fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()
        except ValueError:
            print("Error en el formato de fechas, use MM-DD-AAAA.")
            continue
        break

    fecha_inicio_iso = fecha_inicio.strftime("%Y-%m-%d")
    fecha_fin_iso = fecha_fin.strftime("%Y-%m-%d")

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """SELECT r.id_reservaciones, c.nombre || ' ' || c.apellido AS cliente, r.id_sala, r.fecha, t.turno, r.evento
                   FROM Reservaciones r
                   JOIN Clientes c ON r.id_cliente = c.id_cliente
                   JOIN Salas s ON r.id_sala = s.id_sala
                   JOIN Turnos t ON r.id_turno = t.id_turno
                   WHERE r.estatus = 'Activa'
                   ORDER BY r.fecha ASC"""
            )
            filas = cursor.fetchall()
    except Error as e:
        print(e)
        return

    reservaciones = [
        fila
        for fila in filas
        if fecha_inicio_iso
        <= datetime.datetime.strptime(fila[3], "%Y-%m-%d").strftime("%Y-%m-%d")
        <= fecha_fin_iso
    ]

    if not reservaciones:
        print("No hay reservaciones en el rango indicado.")
        return

    print("*" * 104)
    print(f"**{'RESERVACIONES ENCONTRADAS':^100}** ")
    print("*" * 104)
    print(
        "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
            "ID Reserva",
            "Cliente",
            "ID Sala",
            "Fecha",
            "Turno",
            "Nombre evento",
        )
    )
    print("*" * 104)

    for reserva in reservaciones:
        fecha_evento = datetime.datetime.strptime(reserva[3], "%Y-%m-%d").strftime(
            "%m-%d-%Y"
        )
        print(
            "{:<12} {:<35} {:<10} {:<12} {:<10} {:<25}".format(
                reserva[0],
                reserva[1],
                reserva[2],
                fecha_evento,
                reserva[4],
                reserva[5],
            )
        )
    else:
        print("*" * 104)

    while True:
        entrada = input(
            "Ingrese el ID de la reservación a cancelar (ENTER para salir): "
        ).strip()
        if entrada == "":
            print("Operación cancelada.")
            return
        try:
            id_reservacion = int(entrada)
        except ValueError:
            print("Debe ingresar un número válido.")
            continue
        if id_reservacion not in [r[0] for r in reservaciones]:
            print("Folio no válido. Intente nuevamente.")
            continue
        break
    hoy = datetime.date.today()
    fecha_evento = datetime.datetime.strptime(
        [r[3] for r in reservaciones if r[0] == id_reservacion][0], "%Y-%m-%d"
    ).date()
    if (fecha_evento - hoy).days < 2:
        print("La reservación no puede cancelarse con menos de 2 días de anticipación.")
        return
    confirmacion = (
        input(
            f"Confirma que desea cancelar la reservación {id_reservacion}? Si (S) / No (Cualquier otra tecla.) (Presiona ENTER para cancelar y volver al menu.): "
        )
        .strip()
        .upper()
    )
    if confirmacion != "S":
        print("Operación cancelada por el usuario.")
        return
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "UPDATE Reservaciones SET estatus = 'Cancelada', fecha = NULL, id_turno = NULL WHERE id_reservaciones = ?",
                (id_reservacion,),
            )
        print(f"Reservación {id_reservacion} cancelada exitosamente.")
    except Error as e:
        print(e)


def exportar_csv():
    """Exporta las reservaciones a un archivo CSV"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("""SELECT r.id_reservaciones,
              c.nombre || ' ' || c.apellido AS nombre_cliente,
              s.nombre AS nombre_sala,
              r.fecha,
              t.turno,
              r.evento
       FROM Reservaciones r
       JOIN Clientes c ON r.id_cliente = c.id_cliente
       JOIN Salas s ON r.id_sala = s.id_sala
       JOIN Turnos t ON r.id_turno = t.id_turno
       WHERE r.estatus = 'Activa' 
       ORDER BY r.fecha ASC""")
            reservaciones = cursor.fetchall()
    except Error as e:
        print(e)

    if not reservaciones:
        print("No hay reservaciones para exportar.")
        return

    try:
        with open(
            "Reservaciones.csv", mode="w", encoding="utf-8", newline=""
        ) as archivo:
            writer = csv.writer(archivo)

            encabezados = [
                "Clave".ljust(7),
                "Cliente".ljust(30),
                "Sala".ljust(20),
                "Fecha".ljust(12),
                "Turno".ljust(10),
                "Evento".ljust(25),
            ]
            writer.writerow(encabezados)
            writer.writerow(["-" * 7, "-" * 30, "-" * 20, "-" * 12, "-" * 10, "-" * 25])

            for fila in reservaciones:
                id_reservacion, nombre_cliente, nombre_sala, fecha, turno, evento = fila
                fecha_formateada = datetime.datetime.strptime(
                    fecha, "%Y-%m-%d"
                ).strftime("%m-%d-%Y")
                writer.writerow(
                    [
                        str(id_reservacion).ljust(7),
                        nombre_cliente.ljust(30),
                        nombre_sala.ljust(20),
                        fecha_formateada.ljust(12),
                        turno.ljust(10),
                        evento.ljust(25),
                    ]
                )

        print("Archivo 'reservaciones.csv' creado exitosamente.")
    except Exception as e:
        print(f"Error al crear el archivo {e}")


def exportar_excel():
    """Funcion que exporta las reservaciones a un archivo Excel"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("""SELECT r.id_reservaciones,
              c.nombre || ' ' || c.apellido AS nombre_cliente,
              s.nombre AS nombre_sala,
              r.fecha,
              t.turno,
              r.evento
       FROM Reservaciones r
       JOIN Clientes c ON r.id_cliente = c.id_cliente
       JOIN Salas s ON r.id_sala = s.id_sala
       JOIN Turnos t ON r.id_turno = t.id_turno
       WHERE r.estatus = 'Activa' 
       ORDER BY r.fecha ASC""")
            reservaciones = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not reservaciones:
        print("No hay reservaciones para exportar.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(row=1, column=1)
    cell.value = "REPORTE DE RESERVACIONES"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    encabezados = ["Clave Reservacion", "Cliente", "Sala", "Fecha", "Turno", "Evento"]
    negrita = Font(bold=True)
    borde_grueso = Border(bottom=Side(style="thick"))
    alineado = Alignment(horizontal="center")

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=3, column=col, value=encabezado)
        celda.font = negrita
        celda.alignment = alineado
        celda.border = borde_grueso

    for fila, dato in enumerate(reservaciones, start=4):
        id_reservacion, nombre_cliente, nombre_sala, fecha, turno, evento = dato
        fecha_formateada = datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime(
            "%m-%d-%Y"
        )
        valores = [
            id_reservacion,
            nombre_cliente,
            nombre_sala,
            fecha_formateada,
            turno,
            evento,
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(horizontal="center")

    for col in range(1, 7):
        max_len = max(
            len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)]
        )
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3

    ruta = os.path.join(os.getcwd(), "DatosReservaciones.xlsx")
    wb.save(ruta)
    print("Archivo Excel exportado correctamente")


def exportar_json():
    """Exporta las reservaciones de la base de datos a un archivo JSON."""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("""SELECT r.id_reservaciones,
              c.nombre || ' ' || c.apellido AS nombre_cliente,
              s.nombre AS nombre_sala,
              r.fecha,
              t.turno,
              r.evento
       FROM Reservaciones r
       JOIN Clientes c ON r.id_cliente = c.id_cliente
       JOIN Salas s ON r.id_sala = s.id_sala
       JOIN Turnos t ON r.id_turno = t.id_turno
       WHERE r.estatus = 'Activa' 
       ORDER BY r.fecha ASC""")
            reservaciones = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not reservaciones:
        print("No hay reservaciones para exportar a JSON.")
        return

    lista = []
    for idr, cliente, sala, fecha, turno, evento in reservaciones:
        fecha_formateada = datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime(
            "%m-%d-%Y"
        )
        lista.append(
            {
                "Clave": idr,
                "Cliente": cliente,
                "Sala": sala,
                "Fecha": fecha_formateada,
                "Turno": turno,
                "Evento": evento,
            }
        )
    try:
        ruta = os.path.join(os.getcwd(), "ReservacionesJSON.json")
        with open(ruta, "w", encoding="utf-8") as file:
            json.dump(lista, file, indent=4, ensure_ascii=False)
        print("Archivo JSON exportado correctamente")
    except Exception as e:
        print(f"Error al exportar a JSON: {e}")


def main():
    while True:
        print("\nMENU PRINCIPAL.")
        print("Opciones disponibles: ")
        print("1. Registrar nueva reservacion.")
        print("2. Editar nombre de reservacion.")
        print("3. Consultar reservaciones.")
        print("4. Cancelar reservacion.")
        print("5. Registrar nuevo cliente.")
        print("6. Registrar nueva sala.")
        print("7. Salir.\n")
        opcion = input("Selecciona la opcion que necesites (1-7): ")
        if opcion.isdigit():
            if opcion == "1":
                registrar_reservacion()
            elif opcion == "2":
                editar_reservacion()
            elif opcion == "3":
                consultar_reservacion()
            elif opcion == "4":
                cancelar_reservacion()
            elif opcion == "5":
                registrar_cliente()
            elif opcion == "6":
                registrar_sala()
            elif opcion == "7":
                salir = input(
                    "¿Desea salir realmente? Si (s) / No (Presione cualquier otra tecla): "
                ).lower()
                if salir == "s":
                    print("Saliendo del programa...")
                    break
            else:
                print("Error, ingrese una opción valida")
        else:
            print("Error, ingrese una opción valida")


if __name__ == "__main__":
    iniciar_bd()
    existente = verificar_estado_inicial()
    main()
